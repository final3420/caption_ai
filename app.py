import streamlit as st
import base64
import json
import io
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import requests
from PIL import Image
from openai import OpenAI

try:
    import arabic_reshaper
    from bidi.algorithm import get_display as bidi_display
    def fix_arabic(text: str) -> str:
        if not text:
            return text
        reshaped = arabic_reshaper.reshape(text)
        return bidi_display(reshaped)
except ImportError:
    def fix_arabic(text: str) -> str:
        return text

st.set_page_config(
    page_title="Caption.ai",
    page_icon="✦",
    layout="wide",
    initial_sidebar_state="expanded",
)
import openpyxl
from openpyxl import load_workbook

EXCEL_PATH = "ratings_history.xlsx"

def init_excel():
    """Create Excel file with headers if it doesn't exist."""
    try:
        load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ratings"
        ws.append([
            "Product", "Platform",
            "C_Persuasiveness", "C_Professionalism", "C_Audience_Fit", "C_Creativity", "C_Avg",
            "B_Persuasiveness", "B_Professionalism", "B_Audience_Fit", "B_Creativity", "B_Avg",
            "AI_C_Persuasiveness", "AI_C_Professionalism", "AI_C_Audience_Fit", "AI_C_Creativity", "AI_C_Avg",
            "AI_B_Persuasiveness", "AI_B_Professionalism", "AI_B_Audience_Fit", "AI_B_Creativity", "AI_B_Avg",
            "Caption_Simple", "Caption_Structured", "Timestamp"
        ])
        wb.save(EXCEL_PATH)

@st.cache_data(ttl=60)
def load_ratings_from_excel():
    """Load all ratings from Excel into a DataFrame (cached for 60s)."""
    try:
        return pd.read_excel(EXCEL_PATH)
    except FileNotFoundError:
        return pd.DataFrame()

def save_rating_to_excel(entry, ai_ratings=None):
    """Append a rating row to the Excel file and clear cache."""
    import datetime
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Ratings"]
    simple_scores = entry.get("simple_scores", [0,0,0,0])
    str_scores    = entry.get("str_scores",    [0,0,0,0])
    ai_s  = ai_ratings.get("simple_scores", ["-","-","-","-"]) if ai_ratings else ["-","-","-","-"]
    ai_st = ai_ratings.get("str_scores",    ["-","-","-","-"]) if ai_ratings else ["-","-","-","-"]
    ai_s_avg  = round(sum(ai_s)/len(ai_s), 2)  if ai_ratings and all(isinstance(x, (int,float)) for x in ai_s)  else "-"
    ai_st_avg = round(sum(ai_st)/len(ai_st), 2) if ai_ratings and all(isinstance(x, (int,float)) for x in ai_st) else "-"
    ws.append([
        entry.get("product",""), entry.get("platform",""),
        *simple_scores, entry.get("simple_avg",""),
        *str_scores,    entry.get("str_avg",""),
        *ai_s, ai_s_avg,
        *ai_st, ai_st_avg,
        entry.get("caption_simple",""),
        entry.get("caption_structured",""),
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    wb.save(EXCEL_PATH)
    load_ratings_from_excel.clear()  # clear cache after new save

init_excel()
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=IBM+Plex+Sans+Arabic:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans Arabic', 'Syne', sans-serif; }
.stApp { background: #0D0D0D; color: #F0EDE8; }
[data-testid="stSidebar"] { background: #111111 !important; border-right: 1px solid #222; }
[data-testid="stSidebar"] * { color: #C8C4BC !important; }
[data-testid="stSidebar"] .stTextInput input, [data-testid="stSidebar"] select { background: #1A1A1A !important; border: 1px solid #2A2A2A !important; color: #F0EDE8 !important; border-radius: 8px !important; }
[data-testid="stSidebar"] label { color: #888 !important; font-size: 11px !important; font-weight: 500 !important; letter-spacing: 0.08em !important; text-transform: uppercase !important; }
.brand-header { display: flex; align-items: baseline; gap: 10px; padding: 2rem 0 0.5rem; border-bottom: 1px solid #1E1E1E; margin-bottom: 1.5rem; }
.brand-name { font-family: 'Syne', sans-serif; font-size: 2.6rem; font-weight: 800; color: #F0EDE8; letter-spacing: -0.03em; }
.brand-dot { font-family: 'Syne', sans-serif; font-size: 2.6rem; font-weight: 800; color: #C8F135; }
.brand-tag { font-size: 11px; font-weight: 500; letter-spacing: 0.12em; text-transform: uppercase; color: #555; margin-left: 6px; padding: 3px 10px; border: 1px solid #2A2A2A; border-radius: 20px; }
.brand-sub { font-size: 13px; color: #555; padding: 0.4rem 0 1.5rem; letter-spacing: 0.02em; }
.section-label { font-size: 10px; font-weight: 600; letter-spacing: 0.15em; text-transform: uppercase; color: #555; margin-bottom: 10px; display: flex; align-items: center; gap: 8px; }
.section-label::after { content: ''; flex: 1; height: 1px; background: #1E1E1E; }
[data-testid="stFileUploader"] { background: #111 !important; border: 1px dashed #2A2A2A !important; border-radius: 12px !important; transition: border-color 0.2s; }
[data-testid="stFileUploader"]:hover { border-color: #C8F135 !important; }
[data-testid="stFileUploader"] * { color: #666 !important; }
[data-testid="stFileUploader"] button { background: #1A1A1A !important; color: #C8F135 !important; border: 1px solid #2A2A2A !important; border-radius: 8px !important; }
.stTextInput input, .stSelectbox select, .stTextArea textarea { background: #111 !important; border: 1px solid #222 !important; color: #F0EDE8 !important; border-radius: 8px !important; font-size: 14px !important; }
.stTextInput input:focus, .stTextArea textarea:focus { border-color: #C8F135 !important; box-shadow: 0 0 0 1px #C8F13522 !important; }
.stTextInput label, .stSelectbox label, .stTextArea label { color: #666 !important; font-size: 11px !important; font-weight: 500 !important; letter-spacing: 0.08em !important; text-transform: uppercase !important; }
[data-baseweb="select"] > div { background: #111 !important; border: 1px solid #222 !important; border-radius: 8px !important; color: #F0EDE8 !important; }
[data-baseweb="popover"] { background: #1A1A1A !important; border: 1px solid #333 !important; }
.stButton > button { background: #C8F135 !important; color: #0D0D0D !important; border: none !important; border-radius: 8px !important; font-weight: 700 !important; font-size: 13px !important; letter-spacing: 0.04em !important; padding: 0.6rem 1.4rem !important; transition: all 0.15s !important; }
.stButton > button:hover { background: #D9FF4A !important; transform: translateY(-1px) !important; }
.stButton > button[kind="secondary"] { background: #1A1A1A !important; color: #C8C4BC !important; border: 1px solid #2A2A2A !important; }
.stButton > button[kind="secondary"]:hover { background: #222 !important; border-color: #444 !important; }
.caption-card { background: #111; border-radius: 12px; border: 1px solid #1E1E1E; padding: 1.2rem 1.4rem; font-size: 14px; line-height: 1.85; min-height: 130px; color: #C8C4BC; position: relative; transition: border-color 0.2s; }
.caption-card:hover { border-color: #2A2A2A; }
.caption-card-simple { border-top: 2px solid #F59E0B; }
.caption-card-structured { border-top: 2px solid #C8F135; }
.caption-type-badge { display: inline-flex; align-items: center; gap: 6px; font-size: 10px; font-weight: 600; letter-spacing: 0.12em; text-transform: uppercase; padding: 4px 10px; border-radius: 4px; margin-bottom: 10px; }
.badge-simple { background: #F59E0B18; color: #F59E0B; border: 1px solid #F59E0B33; }
.badge-structured { background: #C8F13518; color: #C8F135; border: 1px solid #C8F13533; }
.badge-ai { background: #818CF818; color: #818CF8; border: 1px solid #818CF833; }
.prompt-box { background: #0A0A0A; border: 1px solid #1A1A1A; border-radius: 10px; padding: 1rem 1.2rem; font-family: 'IBM Plex Mono', monospace; font-size: 12px; color: #666; line-height: 1.7; white-space: pre-wrap; }
.kw-box { background: #C8F13509; border: 1px solid #C8F13522; border-radius: 8px; padding: 0.75rem 1rem; font-size: 13px; color: #9DC43A; line-height: 1.7; margin-top: 6px; }
.metric-card { flex: 1; background: #111; border: 1px solid #1E1E1E; border-radius: 10px; padding: 1rem; text-align: center; }
.metric-label { font-size: 10px; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase; color: #555; margin-bottom: 6px; }
.metric-value { font-family: 'Syne', sans-serif; font-size: 1.8rem; font-weight: 700; color: #F0EDE8; line-height: 1; }
.metric-delta-pos { color: #C8F135; font-size: 12px; margin-top: 4px; }
.metric-delta-neg { color: #F59E0B; font-size: 12px; margin-top: 4px; }
.winner-banner { background: linear-gradient(135deg, #C8F13510, #C8F13520); border: 1px solid #C8F13533; border-radius: 10px; padding: 1rem 1.4rem; display: flex; align-items: center; gap: 12px; color: #C8F135; font-weight: 600; font-size: 14px; margin-top: 1rem; }
.ai-rating-box { background: #0F0F1A; border: 1px solid #818CF833; border-radius: 12px; padding: 1.2rem 1.4rem; margin-top: 1rem; }
.ai-score-pill { display: inline-flex; align-items: center; gap: 6px; background: #818CF818; border: 1px solid #818CF833; border-radius: 20px; padding: 4px 12px; font-size: 12px; color: #818CF8; font-weight: 600; margin: 4px; }
.ai-reasoning { font-size: 12px; color: #555; line-height: 1.7; margin-top: 10px; font-style: italic; border-left: 2px solid #818CF833; padding-left: 10px; }
.stTabs [data-baseweb="tab-list"] { background: transparent !important; border-bottom: 1px solid #1E1E1E !important; gap: 0 !important; }
.stTabs [data-baseweb="tab"] { background: transparent !important; color: #555 !important; border: none !important; font-size: 12px !important; font-weight: 600 !important; letter-spacing: 0.08em !important; text-transform: uppercase !important; padding: 0.75rem 1.2rem !important; }
.stTabs [aria-selected="true"] { color: #F0EDE8 !important; border-bottom: 2px solid #C8F135 !important; }
.stSlider > div > div > div { background: #C8F135 !important; }
.stSlider label { color: #888 !important; font-size: 11px !important; text-transform: uppercase !important; letter-spacing: 0.08em !important; }
hr { border-color: #1A1A1A !important; }
.stExpander { background: #111 !important; border: 1px solid #1E1E1E !important; border-radius: 10px !important; }
.stExpander summary { color: #888 !important; }
.stSuccess { background: #C8F13511 !important; border: 1px solid #C8F13533 !important; color: #9DC43A !important; border-radius: 8px !important; }
.stError { background: #FF444411 !important; border: 1px solid #FF444433 !important; border-radius: 8px !important; }
.stWarning { background: #F59E0B11 !important; border: 1px solid #F59E0B33 !important; border-radius: 8px !important; }
.stInfo { background: #1A1A1A !important; border: 1px solid #2A2A2A !important; border-radius: 8px !important; }
.stDataFrame { border: 1px solid #1E1E1E !important; border-radius: 10px !important; overflow: hidden; }
[data-testid="stImage"] img { border-radius: 12px; border: 1px solid #1E1E1E; }
.rtl { direction: rtl; text-align: right; }
.ltr { direction: ltr; text-align: left; }
.stProgress > div > div { background: #C8F135 !important; }
.how-box { background: #0A0A0A; border: 1px solid #1A1A1A; border-radius: 12px; padding: 1.2rem 1.4rem; margin-bottom: 1.2rem; }
.how-num { font-family: 'Syne', sans-serif; font-size: 2rem; font-weight: 800; color: #C8F135; line-height: 1; margin-bottom: 6px; }
.how-title { font-size: 13px; font-weight: 600; color: #F0EDE8; margin-bottom: 4px; }
.how-desc { font-size: 12px; color: #555; line-height: 1.6; }
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #0D0D0D; }
::-webkit-scrollbar-thumb { background: #2A2A2A; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #3A3A3A; }
</style>
""", unsafe_allow_html=True)

# ─── i18n ───────────────────────────────────────────────────────────────────
T = {
    "en": {
        "lang_label": "Language",
        "apikey_label": "OpenAI API Key",
        "apikey_help": "In-session only — never stored",
        "apikey_placeholder": "sk-...",
        "tab1": "Generate", "tab2": "Analysis", "tab3": "Batch", "tab4": "Export",
        "upload_header": "01 — Upload Product Image",
        "upload_hint": "PNG · JPG · WEBP",
        "params_header": "02 — Marketing Parameters",
        "product_label": "Product Name",
        "audience_label": "Target Audience",
        "tone_label": "Tone",
        "platform_label": "Platform",
        "usp_label": "Unique Selling Point (USP)",
        "kw_header": "03 — SEO Keywords",
        "kw_manual_label": "Enter keywords (comma-separated)",
        "kw_suggest_btn": "Suggest with AI",
        "kw_suggest_spinner": "Generating keywords...",
        "prompts_header": "04 — Prompt Preview",
        "tab_simple": "Simple (C)",
        "tab_str": "Structured (B)",
        "generate_btn": "Generate Both Captions",
        "simple_caption": "Simple Prompt (C)",
        "str_caption": "Structured Prompt (B)",
        "rate_header": "05 — Rate & Compare",
        "criteria": ["Persuasiveness", "Professionalism", "Audience Fit", "Creativity"],
        "rating_btn": "Save Ratings",
        "ai_rate_btn": "⚡ Rate with AI",
        "ai_rate_spinner": "AI is evaluating captions...",
        "ai_rating_header": "AI Evaluation",
        "analysis_header": "Results Analysis",
        "no_ratings": "Complete the generate step and save ratings to see analysis.",
        "export_header": "Export Results",
        "export_btn": "Download JSON",
        "export_csv": "Download CSV",
        "batch_header": "Batch Experiment",
        "batch_add": "Add Product",
        "batch_run": "Run Batch",
        "batch_empty": "Add at least one product to run a batch.",
        "err_nokey": "Please enter your OpenAI API key in the sidebar.",
        "err_noimg": "Please upload a product image.",
        "winner": "Better performing prompt",
        "ai_winner": "AI says better prompt",
        "avg_score": "Avg Score",
        "tones": ["persuasive", "professional", "casual", "luxury", "energetic"],
        "platforms": ["Instagram", "Facebook", "LinkedIn", "Twitter/X", "General"],
        "how_step1_t": "Upload Image", "how_step1_d": "Drop your product photo — any format works",
        "how_step2_t": "Set Parameters", "how_step2_d": "Define audience, tone, platform & USP",
        "how_step3_t": "Add Keywords", "how_step3_d": "Enter SEO keywords or generate with AI",
        "how_step4_t": "Generate & Rate", "how_step4_d": "Compare simple vs structured prompts side by side",
        "simple_label": "Simple prompt uses product name + image only",
        "structured_label": "Structured prompt uses all parameters + SEO keywords",
        "caption_empty": "Caption will appear here after generation",
        "rate_hint_c": "Rate Caption C (Simple)",
        "rate_hint_b": "Rate Caption B (Structured)",
        "sidebar_model": "Model: GPT-4o Vision",
        "sidebar_scale": "Scale: 5-point Likert",
        "sidebar_study": "Prompt Engineering Research",
    },
    "ar": {
        "lang_label": "اللغة",
        "apikey_label": "مفتاح OpenAI API",
        "apikey_help": "يُستخدم في الجلسة فقط — لا يُخزَّن",
        "apikey_placeholder": "sk-...",
        "tab1": "توليد", "tab2": "التحليل", "tab3": "دفعة", "tab4": "تصدير",
        "upload_header": "01 — رفع صورة المنتج",
        "upload_hint": "PNG · JPG · WEBP",
        "params_header": "02 — معاملات التسويق",
        "product_label": "اسم المنتج",
        "audience_label": "الجمهور المستهدف",
        "tone_label": "الأسلوب",
        "platform_label": "المنصة",
        "usp_label": "نقطة البيع الفريدة",
        "kw_header": "03 — الكلمات المفتاحية",
        "kw_manual_label": "أدخل الكلمات المفتاحية (مفصولة بفاصلة)",
        "kw_suggest_btn": "اقتراح بالذكاء الاصطناعي",
        "kw_suggest_spinner": "جارٍ توليد الكلمات المفتاحية...",
        "prompts_header": "04 — معاينة الموجّه",
        "tab_simple": "البسيط (C)",
        "tab_str": "المنظّم (B)",
        "generate_btn": "توليد كلا التعليقَين",
        "simple_caption": "الموجّه البسيط (C)",
        "str_caption": "الموجّه المنظّم (B)",
        "rate_header": "05 — التقييم والمقارنة",
        "criteria": ["الإقناع", "الاحترافية", "توافق الجمهور", "الإبداع"],
        "rating_btn": "حفظ التقييمات",
        "ai_rate_btn": "⚡ تقييم بالذكاء الاصطناعي",
        "ai_rate_spinner": "الذكاء الاصطناعي يقيّم التعليقات...",
        "ai_rating_header": "تقييم الذكاء الاصطناعي",
        "analysis_header": "تحليل النتائج",
        "no_ratings": "أتمم خطوة التوليد واحفظ التقييمات لرؤية التحليل.",
        "export_header": "تصدير النتائج",
        "export_btn": "تنزيل JSON",
        "export_csv": "تنزيل CSV",
        "batch_header": "تجربة دفعية",
        "batch_add": "إضافة منتج",
        "batch_run": "تشغيل الدفعة",
        "batch_empty": "أضف منتجاً واحداً على الأقل.",
        "err_nokey": "يرجى إدخال مفتاح OpenAI API في الشريط الجانبي.",
        "err_noimg": "يرجى رفع صورة المنتج.",
        "winner": "الموجّه الأفضل أداءً",
        "ai_winner": "الأفضل حسب الذكاء الاصطناعي",
        "avg_score": "متوسط الدرجات",
        "tones": ["مقنع", "احترافي", "غير رسمي", "فاخر", "نشيط"],
        "platforms": ["Instagram", "Facebook", "LinkedIn", "Twitter/X", "عام"],
        "how_step1_t": "رفع الصورة", "how_step1_d": "ارفع صورة المنتج — أي تنسيق",
        "how_step2_t": "إعداد المعاملات", "how_step2_d": "حدّد الجمهور والأسلوب والمنصة",
        "how_step3_t": "الكلمات المفتاحية", "how_step3_d": "أدخل كلمات SEO أو اقترحها بالذكاء الاصطناعي",
        "how_step4_t": "التوليد والتقييم", "how_step4_d": "قارن الموجّه البسيط بالمنظّم جنباً إلى جنب",
        "simple_label": "الموجّه البسيط: اسم المنتج + الصورة فقط",
        "structured_label": "الموجّه المنظّم: جميع المعاملات + كلمات SEO",
        "caption_empty": "سيظهر التعليق هنا بعد التوليد",
        "rate_hint_c": "قيّم التعليق C (البسيط)",
        "rate_hint_b": "قيّم التعليق B (المنظّم)",
        "sidebar_model": "النموذج: GPT-4o Vision",
        "sidebar_scale": "المقياس: ليكرت 5 نقاط",
        "sidebar_study": "بحث هندسة الموجّهات",
    }
}

def t(key):
    return T[st.session_state.lang][key]

# ─── Session State ───────────────────────────────────────────────────────────
defaults = {
    "lang": "en",
    "caption_simple": "",
    "caption_structured": "",
    "ratings": {},
    "all_ratings": [],
    "image_b64": None,
    "image_mime": "image/jpeg",
    "suggested_keywords": "",
    "ai_ratings": {},
    "saved_simple_scores": [],
    "saved_str_scores": [],
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─── Helpers ─────────────────────────────────────────────────────────────────
def image_to_b64(img_bytes):
    return base64.b64encode(img_bytes).decode("utf-8")

def build_simple_prompt(product_name, lang):
    name = product_name if product_name else ("the product" if lang == "en" else "المنتج")
    if lang == "ar":
        return f"اكتب تعليقاً تسويقياً جذاباً للمنتج: {name}. اللغة: العربية."
    return f"Write a compelling marketing caption for this product: {name}."

def build_structured_prompt(params, keywords, lang):
    p = params
    usp_line = (f"\nUSP: {p['usp']}" if p.get("usp") else "") if lang == "en" else \
               (f"\nنقطة البيع الفريدة: {p['usp']}" if p.get("usp") else "")
    kw_line = (f"\nSEO Keywords: {keywords}" if keywords.strip() else "") if lang == "en" else \
              (f"\nكلمات SEO: {keywords}" if keywords.strip() else "")
    if lang == "ar":
        return (f"أنت خبير تسويق رقمي. بناءً على الصورة، اكتب تعليقاً تسويقياً مقنعاً.\n\n"
                f"المنتج: {p['product_name']}\nالجمهور: {p['target_audience']}\n"
                f"الأسلوب: {p['tone']}\nالمنصة: {p['platform']}{usp_line}{kw_line}\n\n"
                f"اكتب تعليقاً واحداً مع CTA مناسب للمنصة.")
    return (f"You are a professional digital marketing expert. Based on the image, write a persuasive caption.\n\n"
            f"Product: {p['product_name']}\nAudience: {p['target_audience']}\n"
            f"Tone: {p['tone']}\nPlatform: {p['platform']}{usp_line}{kw_line}\n\n"
            f"Write one effective caption with a clear CTA for the specified platform.")

def generate_caption(client, prompt, b64, mime):
    response = client.chat.completions.create(
        model="gpt-4o", max_tokens=300,
        messages=[{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
            {"type": "text", "text": prompt},
        ]}],
    )
    return response.choices[0].message.content.strip()

def suggest_keywords(client, product_name, lang):
    prompt = (f"Suggest 8-10 SEO marketing keywords for: {product_name}. Comma-separated only, no explanation."
              if lang == "en" else
              f"اقترح 8-10 كلمات مفتاحية SEO للمنتج: {product_name}. مفصولة بفاصلة فقط.")
    resp = client.chat.completions.create(
        model="gpt-4o", max_tokens=100,
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.choices[0].message.content.strip()

def get_ai_ratings(client, caption_simple, caption_structured, criteria, lang):
    """Ask AI to rate both captions on the given criteria, returns scores + reasoning."""
    criteria_str = ", ".join(criteria)
    if lang == "ar":
        prompt = (
            f"أنت محكّم تسويقي متخصص. قيّم التعليقَين التاليَين على كل معيار من 1 إلى 5.\n\n"
            f"التعليق C (البسيط):\n{caption_simple}\n\n"
            f"التعليق B (المنظّم):\n{caption_structured}\n\n"
            f"المعايير: {criteria_str}\n\n"
            f"أجب بـ JSON فقط بهذا الشكل بالضبط بدون أي نص إضافي:\n"
            f'{{"simple": {{"scores": [{{"criterion": "معيار", "score": 4}}], "reasoning": "..."}}, '
            f'"structured": {{"scores": [{{"criterion": "معيار", "score": 4}}], "reasoning": "..."}}}}'
        )
    else:
        prompt = (
            f"You are a professional marketing evaluator. Rate the following two captions on each criterion from 1 to 5.\n\n"
            f"Caption C (Simple):\n{caption_simple}\n\n"
            f"Caption B (Structured):\n{caption_structured}\n\n"
            f"Criteria: {criteria_str}\n\n"
            f"Reply with ONLY a JSON object in exactly this format, no extra text:\n"
            f'{{"simple": {{"scores": [{{"criterion": "criterion_name", "score": 4}}], "reasoning": "..."}}, '
            f'"structured": {{"scores": [{{"criterion": "criterion_name", "score": 4}}], "reasoning": "..."}}}}'
        )

    resp = client.chat.completions.create(
        model="gpt-4o", max_tokens=600,
        messages=[{"role": "user", "content": prompt}],
        response_format={"type": "json_object"},
    )
    data = json.loads(resp.choices[0].message.content.strip())

    def extract_scores(entry):
        score_map = {s["criterion"]: s["score"] for s in entry["scores"]}
        scores = []
        for i, crit in enumerate(criteria):
            if crit in score_map:
                scores.append(score_map[crit])
            else:
                vals = list(score_map.values())
                scores.append(vals[i] if i < len(vals) else 3)
        return scores, entry.get("reasoning", "")

    simple_scores, simple_reasoning = extract_scores(data["simple"])
    str_scores, str_reasoning = extract_scores(data["structured"])
    return {
        "simple_scores": simple_scores,
        "simple_reasoning": simple_reasoning,
        "str_scores": str_scores,
        "str_reasoning": str_reasoning,
    }

def fix_arabic_labels(labels):
    return [fix_arabic(l) for l in labels]

def radar_chart(simple_scores, str_scores, criteria_labels, lang):
    labels = fix_arabic_labels(criteria_labels) if lang == "ar" else criteria_labels
    N = len(labels)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    sv = simple_scores + simple_scores[:1]
    stv = str_scores + str_scores[:1]
    ac = angles + angles[:1]
    fig, ax = plt.subplots(figsize=(4.5, 4.5), subplot_kw=dict(polar=True), facecolor="#111111")
    ax.set_facecolor("#111111")
    leg_s = fix_arabic("بسيط (C)") if lang == "ar" else "Simple (C)"
    leg_st = fix_arabic("منظّم (B)") if lang == "ar" else "Structured (B)"
    ax.plot(ac, sv, "o-", lw=2, color="#F59E0B", label=leg_s)
    ax.fill(ac, sv, alpha=0.15, color="#F59E0B")
    ax.plot(ac, stv, "o-", lw=2, color="#C8F135", label=leg_st)
    ax.fill(ac, stv, alpha=0.15, color="#C8F135")
    ax.set_xticks(angles)
    ax.set_xticklabels(labels, size=9, color="#888")
    ax.set_ylim(0, 5)
    ax.set_yticks([1,2,3,4,5])
    ax.set_yticklabels(["1","2","3","4","5"], size=7, color="#444")
    ax.spines['polar'].set_color("#222")
    ax.grid(color="#1E1E1E", linewidth=0.8)
    ax.legend(loc="upper right", bbox_to_anchor=(1.4, 1.15), fontsize=9,
              facecolor="#1A1A1A", edgecolor="#2A2A2A", labelcolor="#C8C4BC")
    fig.tight_layout()
    return fig

def bar_chart(simple_scores, str_scores, criteria_labels, lang, bar_color_b="#C8F135", bg="#111111", title=""):
    labels = fix_arabic_labels(criteria_labels) if lang == "ar" else criteria_labels
    x = np.arange(len(labels))
    w = 0.35
    fig, ax = plt.subplots(figsize=(6.5, 3.8), facecolor=bg)
    ax.set_facecolor(bg)
    leg_s = fix_arabic("بسيط (C)") if lang == "ar" else "Simple (C)"
    leg_st = fix_arabic("منظّم (B)") if lang == "ar" else "Structured (B)"
    ax.bar(x - w/2, simple_scores, w, label=leg_s, color="#F59E0B", alpha=0.85)
    ax.bar(x + w/2, str_scores, w, label=leg_st, color=bar_color_b, alpha=0.85)
    for xi, v in zip(x - w/2, simple_scores):
        ax.text(xi, v + 0.08, str(v), ha="center", va="bottom", fontsize=8, color="#888")
    for xi, v in zip(x + w/2, str_scores):
        ax.text(xi, v + 0.08, str(v), ha="center", va="bottom", fontsize=8, color=bar_color_b)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=9, color="#888")
    ax.set_ylim(0, 6)
    ax.set_ylabel("Score (1–5)", color="#555", fontsize=10)
    ax.tick_params(colors="#444")
    ax.axhline(3, color="#2A2A2A", ls="--", lw=0.8)
    if title:
        ax.set_title(title, color=bar_color_b, fontsize=10, fontweight="bold")
    ax.legend(fontsize=9, facecolor="#1A1A1A", edgecolor="#2A2A2A", labelcolor="#C8C4BC")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#1E1E1E")
    ax.spines["bottom"].set_color("#1E1E1E")
    fig.tight_layout()
    return fig

# ─── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div style="padding:1.5rem 0 0.5rem;font-family:\'Syne\',sans-serif;font-size:1.4rem;font-weight:800;color:#F0EDE8;letter-spacing:-0.02em;">caption<span style="color:#C8F135;">.</span>ai</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:10px;font-weight:600;letter-spacing:0.12em;text-transform:uppercase;color:#444;margin-bottom:1.5rem;">Prompt Engineering Tool</div>', unsafe_allow_html=True)
    lang_choice = st.radio(t("lang_label"), ["English", "العربية"],
                           index=0 if st.session_state.lang == "en" else 1, horizontal=True)
    st.session_state.lang = "en" if lang_choice == "English" else "ar"
    st.markdown("---")
    api_key = st.text_input(t("apikey_label"), type="password", help=t("apikey_help"), placeholder=t("apikey_placeholder"))
    st.markdown("---")
    st.markdown(f'<div style="font-size:11px;color:#333;line-height:2;">{t("sidebar_model")}<br>{t("sidebar_scale")}<br>{t("sidebar_study")}</div>', unsafe_allow_html=True)

# ─── Brand Header ────────────────────────────────────────────────────────────
rtl_cls = "rtl" if st.session_state.lang == "ar" else "ltr"
st.markdown(f"""
<div class="{rtl_cls}">
  <div class="brand-header">
    <span class="brand-name">caption</span><span class="brand-dot">.</span><span class="brand-name">ai</span>
    <span class="brand-tag">Prompt Engineering Research</span>
  </div>
  <div class="brand-sub">Compare simple vs. structured AI prompts — generate, rate, and export marketing captions.</div>
</div>
""", unsafe_allow_html=True)

with st.expander("How it works — Quick Guide" if st.session_state.lang == "en" else "كيف يعمل — دليل سريع"):
    g1, g2, g3, g4 = st.columns(4)
    for col, num, title_key, desc_key in [
        (g1, "01", "how_step1_t", "how_step1_d"),
        (g2, "02", "how_step2_t", "how_step2_d"),
        (g3, "03", "how_step3_t", "how_step3_d"),
        (g4, "04", "how_step4_t", "how_step4_d"),
    ]:
        with col:
            st.markdown(f'<div class="how-box {rtl_cls}"><div class="how-num">{num}</div><div class="how-title">{t(title_key)}</div><div class="how-desc">{t(desc_key)}</div></div>', unsafe_allow_html=True)

st.markdown("---")
tab1, tab2, tab3, tab4 = st.tabs([t("tab1"), t("tab2"), t("tab3"), t("tab4")])

# ═══════════════════════════════════════════
# TAB 1 — Generate
# ═══════════════════════════════════════════
with tab1:
    col_left, col_right = st.columns([1, 1.55], gap="large")

    with col_left:
        st.markdown(f'<div class="section-label {rtl_cls}">{t("upload_header")}</div>', unsafe_allow_html=True)
        st.caption(t("upload_hint"))
        uploaded = st.file_uploader("", type=["png","jpg","jpeg","webp"], label_visibility="collapsed")
        if uploaded:
            img_bytes = uploaded.read()
            mime_map = {"png":"image/png","jpg":"image/jpeg","jpeg":"image/jpeg","webp":"image/webp"}
            ext = uploaded.name.rsplit(".",1)[-1].lower()
            st.session_state.image_mime = mime_map.get(ext, "image/jpeg")
            st.session_state.image_b64 = image_to_b64(img_bytes)
            img = Image.open(io.BytesIO(img_bytes))
            st.image(img, use_container_width=True)

        st.markdown("---")
        st.markdown(f'<div class="section-label {rtl_cls}">{t("params_header")}</div>', unsafe_allow_html=True)
        product_name = st.text_input(t("product_label"),
            placeholder="e.g. Wireless Headphones" if st.session_state.lang=="en" else "مثال: سماعات لاسلكية")
        target_audience = st.text_input(t("audience_label"),
            placeholder="e.g. Young professionals 25–35" if st.session_state.lang=="en" else "مثال: المهنيون الشباب 25-35")
        c1, c2 = st.columns(2)
        with c1: tone = st.selectbox(t("tone_label"), t("tones"))
        with c2: platform = st.selectbox(t("platform_label"), t("platforms"))
        usp = st.text_input(t("usp_label"),
            placeholder="e.g. 40-hour battery, noise cancellation" if st.session_state.lang=="en" else "مثال: بطارية 40 ساعة، عزل الضوضاء")

        st.markdown("---")
        st.markdown(f'<div class="section-label {rtl_cls}">{t("kw_header")}</div>', unsafe_allow_html=True)
        manual_kw = st.text_input(t("kw_manual_label"), value=st.session_state.suggested_keywords,
            placeholder="keyword1, keyword2, keyword3", key="manual_kw_input")

        if st.button(f"✦ {t('kw_suggest_btn')}", use_container_width=True):
            if not api_key:
                st.error(t("err_nokey"))
            elif not product_name:
                st.warning("Enter a product name first." if st.session_state.lang=="en" else "أدخل اسم المنتج أولاً.")
            else:
                with st.spinner(t("kw_suggest_spinner")):
                    try:
                        kw_client = OpenAI(api_key=api_key)
                        st.session_state.suggested_keywords = suggest_keywords(kw_client, product_name, st.session_state.lang)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

        if st.session_state.suggested_keywords:
            st.markdown(f'<div class="kw-box {rtl_cls}">✦ {st.session_state.suggested_keywords}</div>', unsafe_allow_html=True)

        final_keywords = manual_kw.strip()

    with col_right:
        params = {
            "product_name": product_name or ("the product" if st.session_state.lang=="en" else "المنتج"),
            "target_audience": target_audience or ("potential customers" if st.session_state.lang=="en" else "العملاء المحتملين"),
            "tone": tone,
            "platform": platform,
            "usp": usp,
        }
        simple_prompt = build_simple_prompt(product_name, st.session_state.lang)
        str_prompt = build_structured_prompt(params, final_keywords, st.session_state.lang)

        st.markdown(f'<div class="section-label {rtl_cls}">{t("prompts_header")}</div>', unsafe_allow_html=True)
        ps1, ps2 = st.columns(2)
        with ps1:
            st.markdown(f'<div class="badge-simple caption-type-badge">C — {t("tab_simple")}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="prompt-box {rtl_cls}">{simple_prompt}</div>', unsafe_allow_html=True)
            st.markdown(f'<div style="font-size:10px;color:#333;margin-top:6px;">{t("simple_label")}</div>', unsafe_allow_html=True)
        with ps2:
            st.markdown(f'<div class="badge-structured caption-type-badge">B — {t("tab_str")}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="prompt-box {rtl_cls}">{str_prompt}</div>', unsafe_allow_html=True)
            st.markdown(f'<div style="font-size:10px;color:#333;margin-top:6px;">{t("structured_label")}</div>', unsafe_allow_html=True)

        st.markdown("---")

        if st.button(f"▶  {t('generate_btn')}", use_container_width=True, type="primary"):
            if not api_key:
                st.error(t("err_nokey"))
            elif not st.session_state.image_b64:
                st.error(t("err_noimg"))
            else:
                client = OpenAI(api_key=api_key)
                with st.spinner("Generating Simple Caption (C)..."):
                    try:
                        st.session_state.caption_simple = generate_caption(
                            client, simple_prompt, st.session_state.image_b64, st.session_state.image_mime)
                    except Exception as e:
                        st.error(f"Error: {e}")
                with st.spinner("Generating Structured Caption (B)..."):
                    try:
                        st.session_state.caption_structured = generate_caption(
                            client, str_prompt, st.session_state.image_b64, st.session_state.image_mime)
                    except Exception as e:
                        st.error(f"Error: {e}")
                st.session_state.ai_ratings = {}

        st.markdown("---")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f'<div class="badge-simple caption-type-badge">C — {t("simple_caption")}</div>', unsafe_allow_html=True)
            content = st.session_state.caption_simple or f'<span style="color:#333;font-style:italic;">{t("caption_empty")}</span>'
            st.markdown(f'<div class="caption-card caption-card-simple {rtl_cls}">{content}</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="badge-structured caption-type-badge">B — {t("str_caption")}</div>', unsafe_allow_html=True)
            content2 = st.session_state.caption_structured or f'<span style="color:#333;font-style:italic;">{t("caption_empty")}</span>'
            st.markdown(f'<div class="caption-card caption-card-structured {rtl_cls}">{content2}</div>', unsafe_allow_html=True)

        # ── Rating Section ──
        if st.session_state.caption_simple or st.session_state.caption_structured:
            st.markdown("---")
            st.markdown(f'<div class="section-label {rtl_cls}">{t("rate_header")}</div>', unsafe_allow_html=True)

            criteria = t("criteria")
            simple_scores, str_scores = [], []

            st.markdown(f'<div style="font-size:11px;color:#555;margin-bottom:12px;">{t("rate_hint_c")}</div>', unsafe_allow_html=True)
            r_cols = st.columns(len(criteria))
            for i, (crit, col) in enumerate(zip(criteria, r_cols)):
                with col:
                    st.caption(crit)
                    g = st.slider(f"C-{crit}", 1, 5, 3, key=f"simple_{i}", label_visibility="collapsed")
                    simple_scores.append(g)

            st.markdown(f'<div style="font-size:11px;color:#555;margin-top:8px;margin-bottom:12px;">{t("rate_hint_b")}</div>', unsafe_allow_html=True)
            r_cols2 = st.columns(len(criteria))
            for i, (crit, col) in enumerate(zip(criteria, r_cols2)):
                with col:
                    st.caption(crit)
                    s = st.slider(f"B-{crit}", 1, 5, 3, key=f"str_{i}", label_visibility="collapsed")
                    str_scores.append(s)

            btn_col1, btn_col2 = st.columns(2)
            with btn_col1:
                if st.button(f"✦ {t('rating_btn')}", use_container_width=True):
                    entry = {
                        "product": product_name, "platform": platform,
                        "caption_simple": st.session_state.caption_simple,
                        "caption_structured": st.session_state.caption_structured,
                        "simple_scores": simple_scores, "str_scores": str_scores,
                        "simple_avg": round(sum(simple_scores)/len(simple_scores), 2),
                        "str_avg":    round(sum(str_scores)/len(str_scores), 2),
                    }
                    st.session_state.ratings = entry
                    st.session_state.all_ratings.append(entry)
                    st.session_state.saved_simple_scores = simple_scores.copy()
                    st.session_state.saved_str_scores = str_scores.copy()
                    save_rating_to_excel(entry, st.session_state.ai_ratings or None)
                    st.success("✦ Ratings saved! Head to the Analysis tab." if st.session_state.lang=="en"
                               else "✦ تم حفظ التقييمات! انتقل إلى تبويب التحليل.")

            with btn_col2:
                if st.button(f"{t('ai_rate_btn')}", use_container_width=True):
                    if not api_key:
                        st.error(t("err_nokey"))
                    elif not st.session_state.caption_simple or not st.session_state.caption_structured:
                        st.warning("Generate both captions first." if st.session_state.lang=="en" else "ولّد التعليقَين أولاً.")
                    else:
                        with st.spinner(t("ai_rate_spinner")):
                            try:
                                ai_client = OpenAI(api_key=api_key)
                                result = get_ai_ratings(
                                    ai_client,
                                    st.session_state.caption_simple,
                                    st.session_state.caption_structured,
                                    criteria,
                                    st.session_state.lang,
                                )
                                st.session_state.ai_ratings = result
                                st.rerun()
                            except Exception as e:
                                st.error(f"AI Rating Error: {e}")

            # ── AI Ratings Display Card ──
            if st.session_state.ai_ratings:
                ai_r = st.session_state.ai_ratings
                ai_simple_scores = ai_r["simple_scores"]
                ai_str_scores = ai_r["str_scores"]
                ai_simple_avg = round(sum(ai_simple_scores) / len(ai_simple_scores), 2)
                ai_str_avg = round(sum(ai_str_scores) / len(ai_str_scores), 2)

                st.markdown(f'<div class="ai-rating-box {rtl_cls}">', unsafe_allow_html=True)
                st.markdown(f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;"><span class="badge-ai caption-type-badge" style="margin:0;">⚡ {t("ai_rating_header")}</span><span style="font-size:11px;color:#555;">GPT-4o evaluated both captions</span></div>', unsafe_allow_html=True)

                ai_c1, ai_c2 = st.columns(2)
                with ai_c1:
                    st.markdown(f'<div style="font-size:11px;color:#F59E0B;font-weight:600;margin-bottom:6px;">C — Simple · Avg: {ai_simple_avg}/5</div>', unsafe_allow_html=True)
                    pills_c = "".join([
                        f'<span class="ai-score-pill" style="color:#F59E0B;border-color:#F59E0B33;background:#F59E0B11;">{crit}: {score}/5</span>'
                        for crit, score in zip(criteria, ai_simple_scores)
                    ])
                    st.markdown(f'<div>{pills_c}</div>', unsafe_allow_html=True)
                    if ai_r.get("simple_reasoning"):
                        st.markdown(f'<div class="ai-reasoning {rtl_cls}">{ai_r["simple_reasoning"]}</div>', unsafe_allow_html=True)

                with ai_c2:
                    st.markdown(f'<div style="font-size:11px;color:#818CF8;font-weight:600;margin-bottom:6px;">B — Structured · Avg: {ai_str_avg}/5</div>', unsafe_allow_html=True)
                    pills_b = "".join([
                        f'<span class="ai-score-pill">{crit}: {score}/5</span>'
                        for crit, score in zip(criteria, ai_str_scores)
                    ])
                    st.markdown(f'<div>{pills_b}</div>', unsafe_allow_html=True)
                    if ai_r.get("str_reasoning"):
                        st.markdown(f'<div class="ai-reasoning {rtl_cls}">{ai_r["str_reasoning"]}</div>', unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════
# TAB 2 — Analysis
# ═══════════════════════════════════════════
with tab2:
    st.markdown(f'<div class="section-label {rtl_cls}">{t("analysis_header")}</div>', unsafe_allow_html=True)

    if not st.session_state.ratings:
        st.info(t("no_ratings"))
    else:
        r = st.session_state.ratings
        criteria = t("criteria")
        simple_scores = st.session_state.saved_simple_scores or r["simple_scores"]
        str_scores = st.session_state.saved_str_scores or r["str_scores"]
        lang = st.session_state.lang

        # ── Human metric cards ──
        m_cols = st.columns(len(criteria) + 2)
        for col, label, simp, strd in zip(m_cols[:len(criteria)], criteria, simple_scores, str_scores):
            delta = strd - simp
            delta_cls = "metric-delta-pos" if delta > 0 else ("metric-delta-neg" if delta < 0 else "metric-delta-pos")
            delta_str = f"+{delta}" if delta >= 0 else str(delta)
            with col:
                st.markdown(f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{strd}/5</div><div class="{delta_cls}">{delta_str} vs C</div></div>', unsafe_allow_html=True)
        simple_avg = r["simple_avg"]
        str_avg = r["str_avg"]
        with m_cols[-2]:
            st.markdown(f'<div class="metric-card" style="border-color:#F59E0B33;"><div class="metric-label">Avg C</div><div class="metric-value" style="color:#F59E0B;">{simple_avg}</div></div>', unsafe_allow_html=True)
        with m_cols[-1]:
            st.markdown(f'<div class="metric-card" style="border-color:#C8F13533;"><div class="metric-label">Avg B</div><div class="metric-value" style="color:#C8F135;">{str_avg}</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # ── Human Rating Charts ──
        human_label = "👤 Human Ratings" if lang == "en" else "👤 تقييم البشر"
        st.markdown(f'<div style="font-size:11px;color:#555;font-weight:600;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:12px;">{human_label}</div>', unsafe_allow_html=True)
        ch1, ch2 = st.columns(2)
        with ch1:
            st.pyplot(radar_chart(simple_scores, str_scores, criteria, lang), use_container_width=True)
        with ch2:
            st.pyplot(bar_chart(simple_scores, str_scores, criteria, lang), use_container_width=True)

        # ── AI Rating Charts ──
        if st.session_state.ai_ratings:
            st.markdown("---")
            ai_label = "⚡ AI Ratings" if lang == "en" else "⚡ تقييم الذكاء الاصطناعي"
            st.markdown(f'<div style="font-size:11px;color:#818CF8;font-weight:600;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:12px;">{ai_label}</div>', unsafe_allow_html=True)

            ai_r = st.session_state.ai_ratings
            ai_s = ai_r["simple_scores"]
            ai_st = ai_r["str_scores"]
            ai_simple_avg = round(sum(ai_s) / len(ai_s), 2)
            ai_str_avg = round(sum(ai_st) / len(ai_st), 2)

            # AI metric cards
            ai_m_cols = st.columns(len(criteria) + 2)
            for col, label, simp, strd in zip(ai_m_cols[:len(criteria)], criteria, ai_s, ai_st):
                delta = strd - simp
                delta_cls = "metric-delta-pos" if delta > 0 else ("metric-delta-neg" if delta < 0 else "metric-delta-pos")
                delta_str = f"+{delta}" if delta >= 0 else str(delta)
                with col:
                    st.markdown(f'<div class="metric-card" style="border-color:#818CF833;"><div class="metric-label">{label}</div><div class="metric-value" style="color:#818CF8;">{strd}/5</div><div class="{delta_cls}">{delta_str} vs C</div></div>', unsafe_allow_html=True)
            with ai_m_cols[-2]:
                st.markdown(f'<div class="metric-card" style="border-color:#F59E0B33;"><div class="metric-label">AI Avg C</div><div class="metric-value" style="color:#F59E0B;">{ai_simple_avg}</div></div>', unsafe_allow_html=True)
            with ai_m_cols[-1]:
                st.markdown(f'<div class="metric-card" style="border-color:#818CF833;"><div class="metric-label">AI Avg B</div><div class="metric-value" style="color:#818CF8;">{ai_str_avg}</div></div>', unsafe_allow_html=True)

            ach1, ach2 = st.columns(2)
            with ach1:
                # Radar with purple for B
                labels = fix_arabic_labels(criteria) if lang == "ar" else criteria
                N = len(labels)
                angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
                sv = ai_s + ai_s[:1]
                stv = ai_st + ai_st[:1]
                ac = angles + angles[:1]
                fig_ai_r, ax_r = plt.subplots(figsize=(4.5, 4.5), subplot_kw=dict(polar=True), facecolor="#0F0F1A")
                ax_r.set_facecolor("#0F0F1A")
                leg_s = fix_arabic("بسيط (C)") if lang == "ar" else "Simple (C)"
                leg_st = fix_arabic("منظّم (B)") if lang == "ar" else "Structured (B)"
                ax_r.plot(ac, sv, "o-", lw=2, color="#F59E0B", label=leg_s)
                ax_r.fill(ac, sv, alpha=0.15, color="#F59E0B")
                ax_r.plot(ac, stv, "o-", lw=2, color="#818CF8", label=leg_st)
                ax_r.fill(ac, stv, alpha=0.15, color="#818CF8")
                ax_r.set_xticks(angles)
                ax_r.set_xticklabels(labels, size=9, color="#888")
                ax_r.set_ylim(0, 5)
                ax_r.set_yticks([1,2,3,4,5])
                ax_r.set_yticklabels(["1","2","3","4","5"], size=7, color="#444")
                ax_r.spines['polar'].set_color("#222")
                ax_r.grid(color="#1A1A2A", linewidth=0.8)
                ax_r.set_title("AI Evaluation" if lang=="en" else fix_arabic("تقييم الذكاء الاصطناعي"),
                               color="#818CF8", fontsize=10, fontweight="bold", pad=15)
                ax_r.legend(loc="upper right", bbox_to_anchor=(1.4, 1.15), fontsize=9,
                            facecolor="#111122", edgecolor="#818CF833", labelcolor="#C8C4BC")
                fig_ai_r.tight_layout()
                st.pyplot(fig_ai_r, use_container_width=True)

            with ach2:
                st.pyplot(bar_chart(ai_s, ai_st, criteria, lang,
                                    bar_color_b="#818CF8", bg="#0F0F1A",
                                    title="AI Evaluation" if lang=="en" else fix_arabic("تقييم الذكاء الاصطناعي")),
                          use_container_width=True)

            # AI winner banner
            ai_winner = (("Structured (B)" if lang=="en" else "المنظّم (B)") if ai_str_avg > ai_simple_avg
                        else (("Simple (C)" if lang=="en" else "البسيط (C)") if ai_simple_avg > ai_str_avg
                        else ("Tie" if lang=="en" else "تعادل")))
            ai_diff = abs(ai_str_avg - ai_simple_avg)
            st.markdown(f'<div class="winner-banner {rtl_cls}" style="background:linear-gradient(135deg,#818CF810,#818CF820);border-color:#818CF833;color:#818CF8;"><span style="font-size:1.4rem;">⚡</span><div>{t("ai_winner")}: <strong>{ai_winner}</strong> — advantage: +{ai_diff:.2f} pts</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # Human winner banner
        winner = (("Structured (B)" if lang=="en" else "المنظّم (B)") if str_avg > simple_avg
                  else (("Simple (C)" if lang=="en" else "البسيط (C)") if simple_avg > str_avg
                  else ("Tie" if lang=="en" else "تعادل")))
        diff = abs(str_avg - simple_avg)
        st.markdown(f'<div class="winner-banner {rtl_cls}"><span style="font-size:1.4rem;">✦</span><div>{t("winner")}: <strong>{winner}</strong> — advantage: +{diff:.2f} pts</div></div>', unsafe_allow_html=True)

        if len(st.session_state.all_ratings) > 1:
            st.markdown("---")
            agg_title = "All Sessions" if lang=="en" else "جميع الجلسات"
            st.markdown(f'<div class="section-label {rtl_cls}">{agg_title}</div>', unsafe_allow_html=True)
            rows = []
            for entry in st.session_state.all_ratings:
                rows.append({
                    "Product": entry["product"], "Platform": entry["platform"],
                    "Simple Avg": entry["simple_avg"], "Structured Avg": entry["str_avg"],
                    "Winner": "Structured" if entry["str_avg"] > entry["simple_avg"] else "Simple"
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════
# TAB 3 — Batch
# ═══════════════════════════════════════════
with tab3:
    st.markdown(f'<div class="section-label {rtl_cls}">{t("batch_header")}</div>', unsafe_allow_html=True)
    st.caption("Run multiple products in one go." if st.session_state.lang=="en" else "شغّل عدة منتجات دفعة واحدة.")

    if "batch_products" not in st.session_state:
        st.session_state.batch_products = []

    with st.expander(f"➕  {t('batch_add')}", expanded=True):
        bc1, bc2 = st.columns(2)
        with bc1:
            b_name = st.text_input("Product Name", key="b_name")
            b_audience = st.text_input("Target Audience", key="b_audience")
            b_usp = st.text_input("USP", key="b_usp")
        with bc2:
            b_tone = st.selectbox("Tone", t("tones"), key="b_tone")
            b_platform = st.selectbox("Platform", t("platforms"), key="b_platform")
            b_url = st.text_input("Image URL", key="b_url", placeholder="https://images.unsplash.com/...")
            b_kw = st.text_input("SEO Keywords (optional)", key="b_kw")

        if st.button("Add to Queue"):
            if b_name and b_url:
                st.session_state.batch_products.append({
                    "product_name": b_name, "target_audience": b_audience,
                    "tone": b_tone, "platform": b_platform, "usp": b_usp,
                    "image_url": b_url, "keywords": b_kw,
                })
                st.success(f"Added: {b_name}")
            else:
                st.warning("Product name and image URL required.")

    if st.session_state.batch_products:
        st.markdown(f'<div style="font-size:12px;color:#555;margin:12px 0;">Queue: <strong style="color:#C8F135;">{len(st.session_state.batch_products)}</strong> products</div>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame(st.session_state.batch_products)[["product_name","platform","tone","usp"]],
                     use_container_width=True, hide_index=True)

        if st.button(f"▶  {t('batch_run')}", type="primary", use_container_width=True):
            if not api_key:
                st.error(t("err_nokey"))
            else:
                client = OpenAI(api_key=api_key)
                batch_results = []
                prog = st.progress(0)
                status = st.empty()
                for i, prod in enumerate(st.session_state.batch_products):
                    status.text(f"Processing {i+1}/{len(st.session_state.batch_products)}: {prod['product_name']}...")
                    try:
                        resp = requests.get(prod["image_url"], timeout=10)
                        mime = resp.headers.get("Content-Type","image/jpeg").split(";")[0]
                        b64 = base64.b64encode(resp.content).decode("utf-8")
                        sp = build_simple_prompt(prod["product_name"], st.session_state.lang)
                        stp = build_structured_prompt(prod, prod.get("keywords",""), st.session_state.lang)
                        cap_s = generate_caption(client, sp, b64, mime)
                        cap_st = generate_caption(client, stp, b64, mime)
                        batch_results.append({"Product": prod["product_name"], "Platform": prod["platform"],
                                              "Caption C (Simple)": cap_s, "Caption B (Structured)": cap_st})
                    except Exception as e:
                        batch_results.append({"Product": prod["product_name"], "Platform": prod["platform"],
                                              "Caption C (Simple)": f"ERROR: {e}", "Caption B (Structured)": f"ERROR: {e}"})
                    prog.progress((i+1)/len(st.session_state.batch_products))
                status.text("✦ Batch complete!")
                st.session_state.batch_results = batch_results

        if "batch_results" in st.session_state and st.session_state.batch_results:
            df_batch = pd.DataFrame(st.session_state.batch_results)
            st.dataframe(df_batch, use_container_width=True, hide_index=True)
            csv = df_batch.to_csv(index=False, encoding="utf-8-sig")
            st.download_button("⬇ Download Batch CSV", csv, "batch_captions.csv", "text/csv")
    else:
        st.info(t("batch_empty"))

# ═══════════════════════════════════════════
# TAB 4 — Export
# ═══════════════════════════════════════════
with tab4:
    st.markdown(f'<div class="section-label {rtl_cls}">{t("export_header")}</div>', unsafe_allow_html=True)
    
    df_excel = load_ratings_from_excel()
    if not df_excel.empty:
        st.markdown("### 📊 All Saved Ratings" if st.session_state.lang=="en" else "### 📊 كل التقييمات المحفوظة")
        st.dataframe(df_excel, use_container_width=True, hide_index=True)
        excel_buf = io.BytesIO()
        df_excel.to_excel(excel_buf, index=False, engine="openpyxl")
        st.download_button(
            "⬇ Download Full Excel History",
            data=excel_buf.getvalue(),
            file_name="all_ratings_history.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    if not st.session_state.ratings:
        st.info(t("no_ratings"))
    else:
        r = st.session_state.ratings
        export_data = {
            "product": r.get("product",""), "platform": r.get("platform",""),
            "caption_simple": r.get("caption_simple",""),
            "caption_structured": r.get("caption_structured",""),
            "ratings": {
                "simple": dict(zip(t("criteria"), r["simple_scores"])),
                "structured": dict(zip(t("criteria"), r["str_scores"])),
            },
            "averages": {"simple": r["simple_avg"], "structured": r["str_avg"]},
        }
        if st.session_state.ai_ratings:
            ai_r = st.session_state.ai_ratings
            export_data["ai_ratings"] = {
                "simple": dict(zip(t("criteria"), ai_r["simple_scores"])),
                "structured": dict(zip(t("criteria"), ai_r["str_scores"])),
                "simple_avg": round(sum(ai_r["simple_scores"]) / len(ai_r["simple_scores"]), 2),
                "structured_avg": round(sum(ai_r["str_scores"]) / len(ai_r["str_scores"]), 2),
                "simple_reasoning": ai_r.get("simple_reasoning",""),
                "structured_reasoning": ai_r.get("str_reasoning",""),
            }
        json_str = json.dumps(export_data, ensure_ascii=False, indent=2)
        st.code(json_str, language="json")
        st.download_button(t("export_btn"), data=json_str, file_name="experiment_results.json",
                           mime="application/json", use_container_width=True)

        if st.session_state.all_ratings:
            st.markdown("---")
            rows = []
            for entry in st.session_state.all_ratings:
                rows.append({
                    "Product": entry.get("product",""), "Platform": entry.get("platform",""),
                    **{f"C_{c}": s for c, s in zip(t("criteria"), entry["simple_scores"])},
                    **{f"B_{c}": s for c, s in zip(t("criteria"), entry["str_scores"])},
                    "C_Avg": entry["simple_avg"], "B_Avg": entry["str_avg"],
                })
            df_all = pd.DataFrame(rows)
            csv_all = df_all.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(t("export_csv"), data=csv_all, file_name="all_ratings.csv",
                               mime="text/csv", use_container_width=True)